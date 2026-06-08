"""追加一条预测记录到 decisions.xlsx。

用法:
    python append_decision.py \\
      --title "决策题目" \\
      --sub "子问题 1 | 子问题 2" \\
      --prob "30 | 45" \\
      --deadline "2026-07-01 | 2026-08-15" \\
      --refs "参考类来源" \\
      --devil "魔鬼代言人理由"

文件不存在时自动创建带表头的 Excel。
"""

import argparse
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

DEFAULT_PATH = r"D:\***REMOVED***\10决策系统\decisions.xlsx"
HEADERS = [
    "日期",
    "决策题目",
    "费米化子问题",
    "我的概率",
    "截止日",
    "参考类来源",
    "魔鬼代言人理由",
    "实际结果",
    "反思",
]


def ensure_workbook(path: str):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "decisions"
    ws.append(HEADERS)
    return wb


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--title", required=True)
    parser.add_argument("--sub", required=True, help="多个子问题用 | 分隔")
    parser.add_argument("--prob", required=True, help="多个概率用 | 分隔，与子问题对齐")
    parser.add_argument("--deadline", required=True, help="多个截止日用 | 分隔")
    parser.add_argument("--refs", default="", help="参考类来源")
    parser.add_argument("--devil", default="", help="魔鬼代言人理由")
    parser.add_argument("--path", default=DEFAULT_PATH, help="Excel 文件路径")
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.path), exist_ok=True)
    wb = ensure_workbook(args.path)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")
    subs = [s.strip() for s in args.sub.split("|")]
    probs = [p.strip() for p in args.prob.split("|")]
    deadlines = [d.strip() for d in args.deadline.split("|")]

    if not (len(subs) == len(probs) == len(deadlines)):
        raise SystemExit(
            f"子问题/概率/截止日数量不一致: {len(subs)} / {len(probs)} / {len(deadlines)}"
        )

    for sub, prob, deadline in zip(subs, probs, deadlines):
        ws.append([today, args.title, sub, prob, deadline, args.refs, args.devil, "", ""])

    wb.save(args.path)
    print(f"已写入 {len(subs)} 条记录 -> {args.path}")


if __name__ == "__main__":
    main()
