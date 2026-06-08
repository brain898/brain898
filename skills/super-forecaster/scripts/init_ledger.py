#!/usr/bin/env python3
"""Initialize an empty decisions.xlsx with the fixed header row.

Usage:
    python init_ledger.py --xlsx "D:/***REMOVED***/10决策系统/decisions.xlsx"

If the file already exists and the header is correct, exits without changes.
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


COLUMNS = [
    "日期",
    "决策题目",
    "费米化子问题",
    "我的概率",
    "截止日",
    "参考类来源",
    "魔鬼代言人理由",
    "实际结果",
    "反思",
]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    args = parser.parse_args()

    path = Path(args.xlsx)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        header_ok = all(
            ws.cell(row=1, column=i + 1).value == c for i, c in enumerate(COLUMNS)
        )
        if header_ok:
            print(f"OK exists header_ok file={path}")
            return
        for idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=col)
        wb.save(path)
        print(f"OK exists header_fixed file={path}")
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "decisions"
    for idx, col in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=col)
    wb.save(path)
    print(f"OK created file={path}")


if __name__ == "__main__":
    main()
