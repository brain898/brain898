#!/usr/bin/env python3
"""Append a decision row to decisions.xlsx.

Fixed columns: 日期 / 决策题目 / 费米化子问题 / 我的概率 / 截止日 / 参考类来源 / 魔鬼代言人理由 / 实际结果 / 反思

Usage:
    python append_decision.py --xlsx "D:/***REMOVED***/10决策系统/decisions.xlsx" \
        --date 2026-06-07 \
        --title "..." \
        --subquestions "Q1...|Q2...|Q3..." \
        --probability "" \
        --deadline 2026-08-31 \
        --reference "..." \
        --devil "..." \
        [--result ""] [--reflection ""]

Creates the file with header row if it does not exist.
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


COLUMNS = [
    "日期",
    "决策题目",
    "费米化子问题",
    "我的概率",
    "截止日",
    "参考类来源",
    "魔鬼代言人理由",
    "实际结果",
    "反思",
]


def ensure_workbook(path: Path) -> "Workbook":
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row == 0 or ws.cell(row=1, column=1).value != COLUMNS[0]:
            for idx, col in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=idx, value=col)
        return wb

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "decisions"
    for idx, col in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=col)
    return wb


def append_row(wb: "Workbook", values: dict) -> int:
    ws = wb.active
    row = [values.get(col, "") for col in COLUMNS]
    ws.append(row)
    return ws.max_row


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to decisions.xlsx")
    parser.add_argument("--date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--title", required=True, help="决策题目（步骤 1 拆出的底层真问题）")
    parser.add_argument("--subquestions", required=True, help="费米化子问题，多条用 | 分隔")
    parser.add_argument("--probability", default="", help="我的概率（留空，由用户填）")
    parser.add_argument("--deadline", required=True, help="子问题最远截止日 YYYY-MM-DD")
    parser.add_argument("--reference", required=True, help="参考类来源")
    parser.add_argument("--devil", required=True, help="魔鬼代言人 2 到 3 条反方，用 | 分隔")
    parser.add_argument("--result", default="", help="实际结果（结算时填）")
    parser.add_argument("--reflection", default="", help="反思（结算时填）")
    args = parser.parse_args()

    path = Path(args.xlsx)
    wb = ensure_workbook(path)
    row_no = append_row(wb, {
        "日期": args.date,
        "决策题目": args.title,
        "费米化子问题": args.subquestions,
        "我的概率": args.probability,
        "截止日": args.deadline,
        "参考类来源": args.reference,
        "魔鬼代言人理由": args.devil,
        "实际结果": args.result,
        "反思": args.reflection,
    })
    wb.save(path)
    print(f"OK row={row_no} file={path}")


if __name__ == "__main__":
    main()
