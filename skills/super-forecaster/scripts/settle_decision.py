#!/usr/bin/env python3
"""Settle due predictions and compute forecaster calibration.

This is the soul of super-forecasting: you don't just predict, you come back,
fill in what actually happened, and measure whether your probabilities were
honest. A 70% forecaster should be right ~70% of the time.

Three modes:

  # 1) List predictions whose deadline has passed but aren't settled yet
  python settle_decision.py --xlsx "./decisions.xlsx" --list-due

  # 2) Settle one row: fill 实际结果 (命中/未命中) and 反思
  python settle_decision.py --xlsx "./decisions.xlsx" --settle --row 2 \
      --result yes --reflection "高估了远程实习的任务密度"

  # 3) Compute calibration over all settled rows (Brier score + bucket table)
  python settle_decision.py --xlsx "./decisions.xlsx" --calibrate

Calibration only counts rows where BOTH 我的概率 and 实际结果 are parseable.
"""

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 1-based column positions, must match init_ledger.py / append_decision.py
COL_DATE = 1
COL_TITLE = 2
COL_SUBQ = 3
COL_PROB = 4
COL_DEADLINE = 5
COL_RESULT = 8
COL_REFLECTION = 9

YES_TOKENS = {"yes", "y", "1", "true", "命中", "是", "成", "对"}
NO_TOKENS = {"no", "n", "0", "false", "未命中", "否", "没", "错", "未中"}


def parse_prob(raw) -> float | None:
    """Accept '62%', '62', '0.62' -> 0.62. Return None if unparseable/empty."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    val = float(m.group())
    if "%" in s or val > 1:
        val = val / 100.0
    if val < 0 or val > 1:
        return None
    return val


def parse_outcome(raw) -> int | None:
    """Map result cell to 1 (hit) / 0 (miss) / None (unsettled)."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    if s in YES_TOKENS:
        return 1
    if s in NO_TOKENS:
        return 0
    return None


def parse_date(raw) -> dt.date | None:
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        return raw.date()
    if isinstance(raw, dt.date):
        return raw
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", str(raw))
    if not m:
        return None
    return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))


def load(path: Path):
    if not path.exists():
        print(f"ERROR: ledger not found: {path}", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path)
    return wb, wb.active


def cmd_list_due(ws):
    today = dt.date.today()
    found = 0
    print(f"到期待结算（截止日 <= {today}，且实际结果为空）：")
    for r in range(2, ws.max_row + 1):
        deadline = parse_date(ws.cell(r, COL_DEADLINE).value)
        result = parse_outcome(ws.cell(r, COL_RESULT).value)
        if deadline is None or result is not None:
            continue
        if deadline <= today:
            found += 1
            title = ws.cell(r, COL_TITLE).value
            prob = ws.cell(r, COL_PROB).value
            print(f"  行 {r} | 截止 {deadline} | P={prob or '(未填)'} | {title}")
    if found == 0:
        print("  （无）")
    return found


def cmd_settle(wb, ws, path, row, result, reflection):
    if row < 2 or row > ws.max_row:
        print(f"ERROR: row {row} out of range (2..{ws.max_row})", file=sys.stderr)
        sys.exit(1)
    outcome = parse_outcome(result)
    if outcome is None:
        print(f"ERROR: --result '{result}' 无法解析为命中/未命中", file=sys.stderr)
        sys.exit(1)
    ws.cell(row, COL_RESULT, value="命中" if outcome == 1 else "未命中")
    if reflection is not None:
        ws.cell(row, COL_REFLECTION, value=reflection)
    wb.save(path)
    prob = parse_prob(ws.cell(row, COL_PROB).value)
    hint = ""
    # P=50% 是抛硬币，没有方向可言，跳过方向校准提示。
    # 否则 gap=abs(0.5-outcome) 恒=0.5，命中与否都会被误报「方向判反」。
    if prob is not None and abs(prob - 0.5) > 1e-9:
        gap = abs(prob - outcome)
        if gap > 0.5:
            hint = f"  ⚠ 当初 P={prob:.0%}，结果相反，校准提示：方向判反了"
        elif (prob > 0.5) != (outcome == 1):
            hint = f"  ⚠ 当初 P={prob:.0%}，落在错的一侧"
    print(f"OK 已结算 行={row} 结果={'命中' if outcome else '未命中'}{hint}")


def cmd_calibrate(ws):
    pairs = []
    for r in range(2, ws.max_row + 1):
        p = parse_prob(ws.cell(r, COL_PROB).value)
        o = parse_outcome(ws.cell(r, COL_RESULT).value)
        if p is not None and o is not None:
            pairs.append((p, o))

    n = len(pairs)
    if n == 0:
        print("校准：没有可用样本（需要『我的概率』和『实际结果』都已填）。")
        return

    brier = sum((p - o) ** 2 for p, o in pairs) / n
    base_rate = sum(o for _, o in pairs) / n
    print(f"校准报告（样本 n={n}）")
    print(f"  Brier 分数：{brier:.3f}  （越低越好，0=完美，0.25=瞎猜，参照基准率={base_rate:.0%}）")
    print("  概率分桶 vs 实际命中率：")
    buckets = [(0.0, 0.2), (0.2, 0.4), (0.4, 0.6), (0.6, 0.8), (0.8, 1.01)]
    for lo, hi in buckets:
        sub = [o for p, o in pairs if lo <= p < hi]
        if not sub:
            continue
        hit = sum(sub) / len(sub)
        mid = (lo + min(hi, 1.0)) / 2
        flag = "（自信不足）" if hit > mid + 0.15 else "（过度自信）" if hit < mid - 0.15 else ""
        print(f"    [{lo:.0%}-{min(hi,1.0):.0%}) n={len(sub):>2} 实际命中={hit:.0%} {flag}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--list-due", action="store_true")
    ap.add_argument("--settle", action="store_true")
    ap.add_argument("--calibrate", action="store_true")
    ap.add_argument("--row", type=int)
    ap.add_argument("--result", help="命中/未命中（yes/no/命中/未命中 等）")
    ap.add_argument("--reflection", default=None)
    args = ap.parse_args()

    path = Path(args.xlsx)
    wb, ws = load(path)

    if args.list_due:
        cmd_list_due(ws)
    elif args.settle:
        if args.row is None or args.result is None:
            ap.error("--settle 需要 --row 和 --result")
        cmd_settle(wb, ws, path, args.row, args.result, args.reflection)
    elif args.calibrate:
        cmd_calibrate(ws)
    else:
        ap.error("选一个动作：--list-due / --settle / --calibrate")


if __name__ == "__main__":
    main()
